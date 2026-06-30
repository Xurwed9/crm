import csv
from io import BytesIO

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _build_pdf(title, columns, rows):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 0.3 * inch))

    header_style = ParagraphStyle(
        "Header", parent=styles["Normal"], fontSize=8, textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=7,
    )

    styled_header = [Paragraph(c, header_style) for c in columns]
    table_data = [styled_header]
    for row in rows:
        table_data.append([Paragraph(str(cell or ""), cell_style) for cell in row])

    available_width = landscape(A4)[0] - 72
    col_width = available_width / len(columns)

    table = Table(table_data, colWidths=[col_width] * len(columns), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def _build_excel(title, columns, rows):

    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_csv(columns, rows):

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=\"report.csv\""
    writer = csv.writer(response)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return response


def export_response(data, fmt, filename, columns):

    if fmt == "pdf":
        buffer = _build_pdf(filename.replace("_", " ").title(), columns, data)
        return HttpResponse(buffer, content_type="application/pdf", headers={
            "Content-Disposition": f"attachment; filename=\"{filename}.pdf\"",
        })

    if fmt == "xlsx":
        buffer = _build_excel(filename.replace("_", " ").title(), columns, data)
        return HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename}.xlsx\"",
            },
        )

    if fmt == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f"attachment; filename=\"{filename}.csv\""
        writer = csv.writer(response)
        writer.writerow(columns)
        for row in data:
            writer.writerow(row)
        return response

    return None
